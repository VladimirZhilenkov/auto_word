"""
Service for importing student grades from XLSX files.

Expected XLSX format (based on Moodle-style export):
- Row 1: headers, where column C = 'Адрес электронной почты', column H = 'Оценка/<max>'
- Rows 2..N-1: per-student rows; email in column C, grade in column H
- Row N (last): 'Общее среднее' row — ignored

Stores grade_info on ProgramListener as "{points}/{percent}%".
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..database.connection import DatabaseSession
from ..database.models import Listener, ProgramListener


def _parse_decimal(s: str) -> Optional[float]:
    """Parse Russian-style decimal like '15,00' -> 15.0."""
    if s is None:
        return None
    text = str(s).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _format_grade(points: float) -> str:
    """Format points like 14.0 -> '14,00'."""
    return f"{points:.2f}".replace(".", ",")


class GradesImportService:
    """Parse and import XLSX grade files, matching students by email."""

    def parse_file(self, file_path: str) -> Tuple[Optional[float], Dict[str, float]]:
        """
        Parse the XLSX file.

        Returns:
            (max_grade, {email_lower: points_float})
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Determine max grade from cell H1 ("Оценка/15,00")
        h1 = ws.cell(1, 8).value
        max_grade: Optional[float] = None
        if h1 and isinstance(h1, str) and "/" in h1:
            max_part = h1.split("/", 1)[1]
            max_grade = _parse_decimal(max_part)

        grades: Dict[str, float] = {}
        for row in range(2, ws.max_row + 1):
            last_name = ws.cell(row, 1).value
            email = ws.cell(row, 3).value
            grade_raw = ws.cell(row, 8).value

            # Skip "Общее среднее" summary row
            if isinstance(last_name, str) and "Общее среднее" in last_name:
                continue
            if not email or not isinstance(email, str):
                continue
            grade = _parse_decimal(grade_raw)
            if grade is None:
                continue
            grades[email.strip().lower()] = grade

        return max_grade, grades

    def apply_grades_to_program(
        self,
        program_id: int,
        max_grade: float,
        grades_by_email: Dict[str, float],
    ) -> Dict[str, List[str]]:
        """
        Apply grades to ProgramListener rows for the given program_id.

        Returns a report: {'matched': [...], 'unmatched_emails': [...], 'no_email_in_program': [...]}.
        """
        report: Dict[str, List[str]] = {
            "matched": [],
            "unmatched_emails": [],
            "no_email_in_program": [],
        }

        if max_grade is None or max_grade <= 0:
            raise ValueError("Не удалось определить максимальный балл из файла (ячейка H1)")

        with DatabaseSession() as session:
            assocs = session.query(ProgramListener).filter(
                ProgramListener.program_id == program_id
            ).all()
            for assoc in assocs:
                listener = session.query(Listener).get(assoc.listener_id)
                if not listener:
                    continue
                email = (listener.email or "").strip().lower()
                if not email:
                    report["no_email_in_program"].append(listener.full_name)
                    continue
                points = grades_by_email.get(email)
                if points is None:
                    report["unmatched_emails"].append(f"{listener.full_name} ({email})")
                    continue
                percent = int(round(points / max_grade * 100))
                grade_info = f"{_format_grade(points)}/{percent}%"
                assoc.grade_info = grade_info
                report["matched"].append(f"{listener.full_name}: {grade_info}")

            # Track emails in XLSX that did not correspond to any listener
            matched_emails = set()
            for assoc in assocs:
                listener = session.query(Listener).get(assoc.listener_id)
                if listener and listener.email:
                    matched_emails.add(listener.email.strip().lower())
            for email in grades_by_email:
                if email not in matched_emails:
                    report["unmatched_emails"].append(f"(нет в программе) {email}")

        return report

    def clear_grades_for_program(self, program_id: int) -> int:
        """Remove grade_info from all listeners of a program. Returns count cleared."""
        count = 0
        with DatabaseSession() as session:
            assocs = session.query(ProgramListener).filter(
                ProgramListener.program_id == program_id
            ).all()
            for assoc in assocs:
                if assoc.grade_info:
                    assoc.grade_info = None
                    count += 1
        return count
