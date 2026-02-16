"""
Service for managing Order Journal (журнал регистрации приказов).

Features:
- Independent numbering per journal type (enrollment/admission/graduation)
- Register orders automatically or with custom number
- Query with filters and free-text search
- Update/delete entries with basic validation
- Export to Excel (openpyxl) in required column format
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy import and_, func

from ..database.connection import DatabaseSession
from ..database.models import OrderJournal, Program


ORDER_TYPE_LABELS = {
    'enrollment': 'О зачислении слушателей',
    'admission': 'О допуске слушателей к итоговой аттестации',
    'graduation': 'Об отчислении слушателей',
    'thesis_topics': 'Об утверждении тем итоговых (аттестационных) работ и рецензентов',
    'internship': 'О направлении на стажировку',
    'theory_exam': 'О допуске к сдаче теоретического экзамена',
    'theory_exam_retake': 'О допуске к повторной сдаче теоретического экзамена',
}

REVERSE_ORDER_TYPE_LABELS = {v: k for k, v in ORDER_TYPE_LABELS.items()}


class OrderJournalService:
    """Service layer for OrderJournal model."""

    def get_next_order_number(self, journal_type: str) -> int:
        """Return next sequential number for given journal type."""
        with DatabaseSession() as session:
            max_num = session.query(func.max(OrderJournal.order_number)).filter(
                OrderJournal.journal_type == journal_type
            ).scalar()
            return int(max_num or 0) + 1

    def register_order(
        self,
        journal_type: str,
        title: str,
        program_id: Optional[int] = None,
        program_name: Optional[str] = None,
        executor: str = "",
        order_date: Optional[date] = None,
        notes: Optional[str] = None,
        document_path: Optional[str] = None,
        order_number: Optional[int] = None,
    ) -> int:
        """
        Create and register a new order entry in the journal and return assigned number.
        If order_number is None, assigns next number automatically.
        Validates uniqueness of (journal_type, order_number).
        """
        if order_date is None:
            order_date = date.today()

        with DatabaseSession() as session:
            # Determine number
            assigned_number = order_number or self.get_next_order_number(journal_type)

            # Uniqueness check
            existing = session.query(OrderJournal).filter(
                and_(
                    OrderJournal.journal_type == journal_type,
                    OrderJournal.order_number == assigned_number,
                )
            ).first()
            if existing:
                # If custom number collides, raise error
                raise ValueError(
                    f"Запись с номером №{assigned_number} уже существует для выбранного журнала"
                )

            # Fill program_name from DB if not provided
            resolved_program_name = program_name
            if program_id and not resolved_program_name:
                prog = session.query(Program).get(program_id)
                if prog:
                    resolved_program_name = prog.program_short_name or prog.program_name

            entry = OrderJournal(
                journal_type=journal_type,
                order_number=assigned_number,
                order_date=order_date,
                title=title,
                executor=executor,
                program_id=program_id,
                program_name=resolved_program_name,
                notes=notes,
                document_path=document_path,
            )
            session.add(entry)
            # commit by context
            return assigned_number

    def get_journal_entries(
        self,
        journal_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: Optional[int] = None,
        search: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """Return list of entries as dicts with optional filters and free-text search."""
        with DatabaseSession() as session:
            query = session.query(OrderJournal)
            if journal_type:
                query = query.filter(OrderJournal.journal_type == journal_type)
            if start_date:
                query = query.filter(OrderJournal.order_date >= start_date)
            if end_date:
                query = query.filter(OrderJournal.order_date <= end_date)

            # Order newest on top by number within type, else by date
            query = query.order_by(OrderJournal.order_number.desc(), OrderJournal.order_date.desc())

            rows = query.all() if not limit else query.limit(limit).all()

            results: List[Dict[str, Any]] = []
            for r in rows:
                results.append({
                    'id': r.id,
                    'journal_type': r.journal_type,
                    'order_number': r.order_number,
                    'order_date': r.order_date,
                    'title': r.title,
                    'executor': r.executor,
                    'program_id': r.program_id,
                    'program_name': r.program_name,
                    'notes': r.notes,
                    'created_at': r.created_at,
                    'document_path': r.document_path,
                })

            if search:
                s = (search or '').strip().lower()
                if s:
                    def matches(e: Dict[str, Any]) -> bool:
                        fields = [
                            str(e.get('order_number', '')),
                            str(e.get('title', '') or ''),
                            str(e.get('executor', '') or ''),
                            str(e.get('program_name', '') or ''),
                            str(e.get('notes', '') or ''),
                        ]
                        return any(s in f.lower() for f in fields)
                    results = [e for e in results if matches(e)]

            return results

    def export_journal_to_excel(
        self,
        journal_type: str,
        output_path: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        search: Optional[str] = None,
    ) -> bool:
        """Export filtered journal entries to an Excel file using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            entries = self.get_journal_entries(
                journal_type=journal_type,
                start_date=start_date,
                end_date=end_date,
                search=search,
            )
            if not entries:
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = 'Журнал'

            # Title row
            type_label = ORDER_TYPE_LABELS.get(journal_type, 'Журнал приказов')
            ws.merge_cells('A1:G1')
            title_cell = ws.cell(row=1, column=1, value=f'Журнал приказов директора по личному составу слушателей')
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:G2')
            subtitle_cell = ws.cell(row=2, column=1, value=f'Тип: {type_label}')
            subtitle_cell.font = Font(bold=True, size=11)
            subtitle_cell.alignment = Alignment(horizontal='center')

            headers = ['№ п/п', 'Дата приказа', 'Номер приказа', 'Наименование (краткое содержание)', 'Программа', 'Ответственный исполнитель', 'Примечание']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                from openpyxl.styles import Border, Side
                thin = Side(style='thin')
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Data rows
            for row_idx, e in enumerate(entries, start=5):
                row_data = [
                    row_idx - 4,
                    e['order_date'].strftime('%d.%m.%Y') if e['order_date'] else '',
                    e['order_number'],
                    e['title'] or '',
                    e['program_name'] or '',
                    e['executor'] or '',
                    e['notes'] or '',
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(wrap_text=True)
                    from openpyxl.styles import Border, Side
                    thin = Side(style='thin')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Column widths
            widths = [8, 14, 14, 50, 30, 25, 30]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False

    def update_order_entry(self, order_id: int, **kwargs) -> bool:
        """Update an existing entry by id. Validates uniqueness when changing number/type."""
        allowed = {
            'journal_type', 'order_number', 'order_date', 'title', 'executor',
            'program_id', 'program_name', 'notes', 'document_path'
        }
        fields = {k: v for k, v in kwargs.items() if k in allowed}
        if not fields:
            return False

        with DatabaseSession() as session:
            entry = session.query(OrderJournal).get(order_id)
            if not entry:
                return False

            new_type = fields.get('journal_type', entry.journal_type)
            new_num = fields.get('order_number', entry.order_number)

            # Uniqueness check if changed
            if new_type != entry.journal_type or new_num != entry.order_number:
                exists = session.query(OrderJournal).filter(
                    and_(
                        OrderJournal.journal_type == new_type,
                        OrderJournal.order_number == new_num,
                        OrderJournal.id != order_id,
                    )
                ).first()
                if exists:
                    raise ValueError(
                        f"Запись с номером №{new_num} уже существует для выбранного журнала"
                    )

            # Resolve program_name if program_id provided and name missing
            if 'program_id' in fields and fields.get('program_id') and not fields.get('program_name'):
                prog = session.query(Program).get(fields['program_id'])
                if prog:
                    fields['program_name'] = prog.program_short_name or prog.program_name

            for k, v in fields.items():
                setattr(entry, k, v)
            return True

    def delete_order_entry(self, order_id: int) -> bool:
        """
        Delete entry by id.
        """
        with DatabaseSession() as session:
            entry = session.query(OrderJournal).get(order_id)
            if not entry:
                return False

            session.delete(entry)
            return True
