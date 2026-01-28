"""
Service for managing contract journal entries and generating contract documents.
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from docxtpl import DocxTemplate
from sqlalchemy import and_, func

from ..database.connection import DatabaseSession
from ..database.models import ContractJournal, Listener, Program
from .declension import get_declension_service


def _number_to_words(number: Union[int, float, Decimal, str, None], currency: str = "rub", skip_cents: bool = False) -> str:
    """
    Конвертирует число в сумму прописью на русском языке.
    
    Args:
        number: Число для конвертации
        currency: Валюта ('rub' - рубли, 'usd' - доллары, 'eur' - евро, 'none' - без валюты)
        skip_cents: Если True, не выводить копейки/центы
    
    Returns:
        Строка с суммой прописью
    """
    if number is None:
        return ""
    
    try:
        # Преобразуем в Decimal для точности
        if isinstance(number, str):
            number = number.replace(" ", "").replace(",", ".")
        num = Decimal(str(number))
    except:
        return ""
    
    # Единицы
    units = [
        "", "один", "два", "три", "четыре", "пять", 
        "шесть", "семь", "восемь", "девять"
    ]
    units_female = [
        "", "одна", "две", "три", "четыре", "пять",
        "шесть", "семь", "восемь", "девять"
    ]
    
    # Числа 10-19
    teens = [
        "десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать",
        "пятнадцать", "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"
    ]
    
    # Десятки
    tens = [
        "", "", "двадцать", "тридцать", "сорок", "пятьдесят",
        "шестьдесят", "семьдесят", "восемьдесят", "девяносто"
    ]
    
    # Сотни
    hundreds = [
        "", "сто", "двести", "триста", "четыреста", "пятьсот",
        "шестьсот", "семьсот", "восемьсот", "девятьсот"
    ]
    
    def _get_form(n: int, forms: tuple) -> str:
        """Возвращает правильную форму слова в зависимости от числа."""
        n = abs(n) % 100
        if 11 <= n <= 19:
            return forms[2]
        n = n % 10
        if n == 1:
            return forms[0]
        if 2 <= n <= 4:
            return forms[1]
        return forms[2]
    
    def _convert_group(n: int, female: bool = False) -> str:
        """Конвертирует группу из трёх цифр."""
        result = []
        
        if n >= 100:
            result.append(hundreds[n // 100])
            n %= 100
        
        if 10 <= n <= 19:
            result.append(teens[n - 10])
        else:
            if n >= 10:
                result.append(tens[n // 10])
                n %= 10
            if n > 0:
                if female:
                    result.append(units_female[n])
                else:
                    result.append(units[n])
        
        return " ".join(filter(None, result))
    
    # Разделяем целую и дробную части
    integer_part = int(num)
    fractional_part = int(round((num - integer_part) * 100))
    
    if integer_part == 0:
        result = "ноль"
    else:
        groups = []
        temp = integer_part
        
        # Единицы (рубли - мужской род)
        group = temp % 1000
        if group:
            groups.append(_convert_group(group, female=False))
        temp //= 1000
        
        # Тысячи (женский род)
        group = temp % 1000
        if group:
            word = _get_form(group, ("тысяча", "тысячи", "тысяч"))
            groups.append(_convert_group(group, female=True) + " " + word)
        temp //= 1000
        
        # Миллионы
        group = temp % 1000
        if group:
            word = _get_form(group, ("миллион", "миллиона", "миллионов"))
            groups.append(_convert_group(group, female=False) + " " + word)
        temp //= 1000
        
        # Миллиарды
        group = temp % 1000
        if group:
            word = _get_form(group, ("миллиард", "миллиарда", "миллиардов"))
            groups.append(_convert_group(group, female=False) + " " + word)
        
        result = " ".join(reversed(groups))
    
    # Добавляем валюту
    if currency == "rub":
        rub_word = _get_form(integer_part, ("рубль", "рубля", "рублей"))
        if skip_cents:
            result = f"{result} {rub_word}"
        else:
            kop_word = _get_form(fractional_part, ("копейка", "копейки", "копеек"))
            result = f"{result} {rub_word} {fractional_part:02d} {kop_word}"
    elif currency == "usd":
        usd_word = _get_form(integer_part, ("доллар", "доллара", "долларов"))
        if skip_cents:
            result = f"{result} {usd_word}"
        else:
            cent_word = _get_form(fractional_part, ("цент", "цента", "центов"))
            result = f"{result} {usd_word} {fractional_part:02d} {cent_word}"
    elif currency == "eur":
        eur_word = _get_form(integer_part, ("евро", "евро", "евро"))
        if skip_cents:
            result = f"{result} {eur_word}"
        else:
            cent_word = _get_form(fractional_part, ("цент", "цента", "центов"))
            result = f"{result} {eur_word} {fractional_part:02d} {cent_word}"
    # currency == "none" - без валюты
    
    return result.strip()


def _get_app_dir() -> Path:
    """Return application root directory (works for frozen builds)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent.parent.parent


class ContractJournalService:
    """Service layer for ContractJournal with document generation helpers."""

    DEFAULT_TEMPLATE = "contract_template.docx"
    CONTRACTS_TEMPLATES_SUBDIR = "contracts"

    def __init__(self, templates_dir: Optional[Path] = None, output_dir: Optional[Path] = None):
        app_dir = _get_app_dir()
        base_templates = Path(templates_dir) if templates_dir else app_dir / "templates"
        self.templates_dir = base_templates / self.CONTRACTS_TEMPLATES_SUBDIR
        self.output_dir = Path(output_dir) if output_dir else app_dir / "docx_files" / "contracts"
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.declension = get_declension_service()

    def get_available_templates(self) -> List[str]:
        """Return list of available contract template files."""
        if not self.templates_dir.exists():
            return []
        templates = []
        for pattern in ["*.docx", "*.DOCX"]:
            templates.extend(self.templates_dir.glob(pattern))
        return sorted([t.name for t in templates if not t.name.startswith(("~$", "._"))])

    # ------------------------------------------------------------------
    # Core CRUD
    # ------------------------------------------------------------------
    def get_next_contract_number(self) -> int:
        with DatabaseSession() as session:
            max_num = session.query(func.max(ContractJournal.contract_number)).scalar()
            return int(max_num or 0) + 1

    def _get_next_in_session(self, session) -> int:
        max_num = session.query(func.max(ContractJournal.contract_number)).scalar()
        return int(max_num or 0) + 1

    def create_contract(
        self,
        listener_id: int,
        program_id: Optional[int],
        contract_date: date,
        **kwargs,
    ) -> int:
        with DatabaseSession() as session:
            return self._create_contract_in_session(session, listener_id, program_id, contract_date, **kwargs)

    def _create_contract_in_session(self, session, listener_id: int, program_id: Optional[int], contract_date: date, **kwargs) -> int:
        contract_number = kwargs.get("contract_number")
        if contract_number is None:
            contract_number = self._get_next_in_session(session)
        else:
            exists = session.query(ContractJournal).filter(
                ContractJournal.contract_number == contract_number
            ).first()
            if exists:
                raise ValueError(f"Договор с номером {contract_number} уже существует")

        listener = session.query(Listener).get(listener_id)
        program = session.query(Program).get(program_id) if program_id else None
        if not listener:
            raise ValueError("Слушатель не найден")

        listener_full_name = listener.full_name
        program_name = None
        if program:
            program_name = program.program_short_name or program.program_name

        entry = ContractJournal(
            contract_number=contract_number,
            contract_date=contract_date,
            listener_id=listener_id,
            program_id=program_id,
            listener_full_name=listener_full_name,
            program_name=program_name,
            contract_sum=kwargs.get("contract_sum"),
            payment_type=kwargs.get("payment_type"),
            notes=kwargs.get("notes"),
            document_path=kwargs.get("document_path"),
        )
        session.add(entry)
        return contract_number

    def create_contracts_batch(
        self,
        listener_ids: List[int],
        program_id: Optional[int],
        contract_date: date,
        template_name: Optional[str] = None,
        **kwargs,
    ) -> List[int]:
        created_numbers: List[int] = []
        with DatabaseSession() as session:
            program = session.query(Program).get(program_id) if program_id else None
            next_number = self._get_next_in_session(session)

            for offset, listener_id in enumerate(listener_ids):
                listener = session.query(Listener).get(listener_id)
                if not listener:
                    continue

                current_number = next_number + offset
                document_path = self._render_contract_document(
                    listener=listener,
                    program=program,
                    contract_number=current_number,
                    contract_date=contract_date,
                    template_name=template_name,
                    extra_context=kwargs.get("context") or {},
                    contract_sum=kwargs.get("contract_sum"),
                )

                entry = ContractJournal(
                    contract_number=current_number,
                    contract_date=contract_date,
                    listener_id=listener.id,
                    program_id=program.id if program else None,
                    listener_full_name=listener.full_name,
                    program_name=program.program_short_name or program.program_name if program else None,
                    contract_sum=kwargs.get("contract_sum"),
                    payment_type=kwargs.get("payment_type"),
                    notes=kwargs.get("notes"),
                    document_path=document_path,
                )
                session.add(entry)
                created_numbers.append(current_number)

        return created_numbers

    def get_contracts(
        self,
        program_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        search: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        with DatabaseSession() as session:
            query = session.query(ContractJournal)
            if program_id:
                query = query.filter(ContractJournal.program_id == program_id)
            if start_date:
                query = query.filter(ContractJournal.contract_date >= start_date)
            if end_date:
                query = query.filter(ContractJournal.contract_date <= end_date)

            query = query.order_by(ContractJournal.contract_number.desc())
            rows = query.all()

            items: List[Dict[str, Any]] = []
            for r in rows:
                items.append(
                    {
                        "id": r.id,
                        "contract_number": r.contract_number,
                        "contract_date": r.contract_date,
                        "listener_full_name": r.listener_full_name,
                        "program_name": r.program_name,
                        "contract_sum": r.contract_sum,
                        "payment_type": r.payment_type,
                        "notes": r.notes,
                        "document_path": r.document_path,
                        "listener_id": r.listener_id,
                        "program_id": r.program_id,
                        "created_at": r.created_at,
                    }
                )

            if search:
                s = search.strip().lower()
                items = [
                    i
                    for i in items
                    if s in str(i.get("contract_number", "")).lower()
                    or s in (i.get("listener_full_name") or "").lower()
                    or s in (i.get("program_name") or "").lower()
                    or s in (i.get("notes") or "").lower()
                ]

            return items

    def update_contract(self, contract_id: int, **kwargs) -> bool:
        allowed = {
            "contract_number",
            "contract_date",
            "listener_id",
            "program_id",
            "listener_full_name",
            "program_name",
            "contract_sum",
            "payment_type",
            "notes",
            "document_path",
        }
        fields = {k: v for k, v in kwargs.items() if k in allowed}
        if not fields:
            return False

        with DatabaseSession() as session:
            entry: ContractJournal = session.query(ContractJournal).get(contract_id)
            if not entry:
                return False

            new_number = fields.get("contract_number", entry.contract_number)
            if new_number != entry.contract_number:
                exists = session.query(ContractJournal).filter(
                    and_(ContractJournal.contract_number == new_number, ContractJournal.id != contract_id)
                ).first()
                if exists:
                    raise ValueError(f"Договор с номером {new_number} уже существует")

            if "listener_id" in fields:
                listener = session.query(Listener).get(fields["listener_id"])
                if listener:
                    fields.setdefault("listener_full_name", listener.full_name)
            if "program_id" in fields:
                program = session.query(Program).get(fields["program_id"])
                if program:
                    fields.setdefault("program_name", program.program_short_name or program.program_name)

            for k, v in fields.items():
                setattr(entry, k, v)
            return True

    def delete_contract(self, contract_id: int) -> bool:
        with DatabaseSession() as session:
            entry = session.query(ContractJournal).get(contract_id)
            if not entry:
                return False
            session.delete(entry)
            return True

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def export_to_excel(self, output_path: str, filters: Dict[str, Any]) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter

            entries = self.get_contracts(
                program_id=filters.get("program_id"),
                start_date=filters.get("start_date"),
                end_date=filters.get("end_date"),
                search=filters.get("search"),
            )
            if not entries:
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = "Журнал договоров"

            headers = [
                "№ договора",
                "Дата",
                "ФИО слушателя",
                "Программа",
                "Сумма",
                "Тип оплаты",
                "Примечания",
            ]
            ws.append(headers)

            for idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for item in entries:
                ws.append(
                    [
                        item.get("contract_number"),
                        item.get("contract_date").strftime("%d.%m.%Y") if item.get("contract_date") else "",
                        item.get("listener_full_name") or "",
                        item.get("program_name") or "",
                        item.get("contract_sum") or "",
                        item.get("payment_type") or "",
                        item.get("notes") or "",
                    ]
                )

            widths = [12, 12, 32, 32, 12, 14, 40]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            return True
        except Exception as exc:
            print(f"Excel export failed: {exc}")
            return False

    # ------------------------------------------------------------------
    # Document rendering
    # ------------------------------------------------------------------
    def _render_contract_document(
        self,
        listener: Listener,
        program: Optional[Program],
        contract_number: int,
        contract_date: date,
        template_name: Optional[str] = None,
        extra_context: Optional[Dict[str, Any]] = None,
        contract_sum: Optional[Union[int, float, Decimal, str]] = None,
    ) -> Optional[str]:
        tpl_name = template_name or self.DEFAULT_TEMPLATE
        template_path = self.templates_dir / tpl_name
        if not template_path.exists():
            return None

        context = self._build_context(listener, program, contract_number, contract_date, contract_sum)
        if extra_context:
            context.update(extra_context)

        try:
            doc = DocxTemplate(template_path)
            doc.render(context)
            safe_name = listener.full_name.replace(" ", "_")
            filename = f"Договор_{contract_number:03d}_{safe_name}.docx"
            output_path = self.output_dir / filename
            doc.save(output_path)
            return str(output_path)
        except Exception as exc:
            print(f"Contract render failed: {exc}")
            return None

    def _build_context(
        self,
        listener: Listener,
        program: Optional[Program],
        contract_number: int,
        contract_date: date,
        contract_sum: Optional[Union[int, float, Decimal, str]] = None,
    ) -> Dict[str, Any]:
        # Базовые данные слушателя (без префикса - для совместимости с шаблонами)
        
        # Формируем инициалы: И.О. Фамилия
        first_initial = (listener.first_name[0] + ".") if listener.first_name else ""
        middle_initial = (listener.middle_name[0] + ".") if listener.middle_name else ""
        initials_last_name = f"{first_initial}{middle_initial} {listener.last_name or ''}".strip()
        # Фамилия И.О.
        last_name_initials = f"{listener.last_name or ''} {first_initial}{middle_initial}".strip()
        
        ctx: Dict[str, Any] = {
            # Номер и дата договора
            "contract_number": contract_number,
            "contract_date": contract_date.strftime("%d.%m.%Y"),
            "contract_year": contract_date.strftime("%Y"),
            
            # ФИО слушателя (без префикса)
            "full_name": listener.full_name,
            "last_name": listener.last_name or "",
            "first_name": listener.first_name or "",
            "middle_name": listener.middle_name or "",
            "initials_last_name": initials_last_name,  # И.О. Фамилия
            "last_name_initials": last_name_initials,  # Фамилия И.О.
            
            # Дата рождения
            "birth_date": listener.birth_date.strftime("%d.%m.%Y") if listener.birth_date else "",
            
            # Рабочая информация
            "position": listener.position or "",
            "workplace": listener.workplace or "",
            "region": listener.region or "",
            
            # Контактная информация
            "mobile_phone": listener.mobile_phone or "",
            "work_phone": listener.work_phone or "",
            "email": listener.email or "",
            
            # Паспортные данные
            "passport_series_number": listener.passport_series_number or "",
            "passport_issue_date": listener.passport_issue_date.strftime("%d.%m.%Y") if listener.passport_issue_date else "",
            "passport_issued_by": listener.passport_issued_by or "",
            "passport_department_code": listener.passport_department_code or "",
            
            # Адреса
            "registration_address": listener.registration_address or "",
            "actual_address": listener.actual_address or "",
            
            # Идентификационные данные
            "snils": listener.snils or "",
            "inn": listener.inn or "",
        }
        
        # Дублируем с префиксом listener_ для обратной совместимости
        ctx.update({
            "listener_full_name": listener.full_name,
            "listener_last_name": listener.last_name or "",
            "listener_first_name": listener.first_name or "",
            "listener_middle_name": listener.middle_name or "",
            "listener_initials_last_name": initials_last_name,  # И.О. Фамилия
            "listener_last_name_initials": last_name_initials,  # Фамилия И.О.
            "listener_birth_date": listener.birth_date.strftime("%d.%m.%Y") if listener.birth_date else "",
            "listener_position": listener.position or "",
            "listener_workplace": listener.workplace or "",
            "listener_region": listener.region or "",
            "listener_phone": listener.mobile_phone or "",
            "listener_mobile_phone": listener.mobile_phone or "",
            "listener_work_phone": listener.work_phone or "",
            "listener_email": listener.email or "",
            "listener_passport": listener.passport_series_number or "",
            "listener_passport_series_number": listener.passport_series_number or "",
            "listener_passport_issue_date": listener.passport_issue_date.strftime("%d.%m.%Y") if listener.passport_issue_date else "",
            "listener_passport_issued_by": listener.passport_issued_by or "",
            "listener_passport_department_code": listener.passport_department_code or "",
            "listener_address": listener.registration_address or "",
            "listener_registration_address": listener.registration_address or "",
            "listener_actual_address": listener.actual_address or "",
            "listener_snils": listener.snils or "",
            "listener_inn": listener.inn or "",
        })
        
        # Сумма договора
        if contract_sum is not None:
            ctx["contract_sum"] = contract_sum
            ctx["contract_sum_words"] = _number_to_words(contract_sum, currency="none")
            ctx["contract_sum_words_caps"] = _number_to_words(contract_sum, currency="none").capitalize()
        else:
            ctx["contract_sum"] = ""
            ctx["contract_sum_words"] = ""
            ctx["contract_sum_words_caps"] = ""

        # Образование слушателя
        ctx["listener_education"] = listener.education or ""
        ctx["education"] = listener.education or ""

        # Склонения ФИО (с учётом пола)
        gender = getattr(listener, 'gender', None) or 'M'
        decl = self.declension.get_all_declensions(
            listener.last_name or "",
            listener.first_name or "",
            listener.middle_name,
            gender=gender,
        )
        ctx["listener_full_name_genitive"] = decl.get("full_name_genitive", "")
        ctx["full_name_genitive"] = decl.get("full_name_genitive", "")
        ctx.update(decl)

        # Программа
        if program:
            ctx.update(
                {
                    "program_name": program.program_name or "",
                    "program_short_name": program.program_short_name or program.program_name or "",
                    "program_volume": program.program_volume or "",
                    "training_period": program.training_period or "",
                    "training_duration": program.training_duration or "",
                    "education_form": program.education_form or "",
                    "education_format": program.education_format or "",
                    "expulsion_date": program.expulsion_date.strftime("%d.%m.%Y") if program.expulsion_date else "",
                }
            )
        else:
            ctx.update({
                "program_name": "", 
                "program_short_name": "",
                "program_volume": "", 
                "training_period": "",
                "training_duration": "",
                "education_form": "",
                "education_format": "",
                "expulsion_date": "",
            })

        return ctx

